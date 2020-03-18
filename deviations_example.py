# DO NOT EDIT THIS SCRIPT!#
import os
from configparser import ConfigParser
import winreg
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color, colors
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import subprocess
sys.path.append(r'\Modules')
from LoggerManager import Logger
from itertools import islice

#gets the base location
def get_location():
	if 'Rework' in location:
		return location[:-18]
	else:
		return location

#NOT USED!!
def excel_style(col):
	LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
	""" Convert given row and Column number to an Excel-style cell name. """
	result = []
	while col:
		col, rem = divmod(col - 1, 26)
		result[:0] = LETTERS[rem]
	return ''.join(result)

#gets a list of all the csv files in the current location
def get_csv_files():
	return [file for file in os.listdir(location) if file.endswith('.csv') and not file.startswith('Ω')]

#parse all csv files in current location and combine selected columns ('Dev'/'Actual')
def get_deviations(write_to_file=True):

	#construct dataframe and set first column to the elements list
	#found in first column of first csv file in the location
	deviations = pd.DataFrame()
	deviations['Element'] = pd.read_csv(os.path.join(location, csvfiles[0]))['Element']

	#iterates the Element column and fixes the names to match the master
	row = 0
	for _ in deviations['Element']:
		deviations.at[row, 'Element'] = deviations.at[row, 'Element'].replace(deviations.at[row, 'Element'].split(" ")[0],'')
		row += 1

	#combine 'Dev' column from all csv's in location and export resulting csv
	if cfg.getboolean('ScrapeData', 'deviations'):
		for file in csvfiles:
			currentDF = pd.read_csv(os.path.join(location, file))
			deviations[str(file[:-4])] = currentDF['Dev']
		deviations.to_csv(location + r'\ΩDeviations.csv', index=False)

	#combine 'Actual' column from all csv's and export resulting csv
	#if substitue_dev is True, any row 'Actual' column is null/None replace 
	#with value in 'Dev' column for that row
	if cfg.getboolean('ScrapeData', 'actuals'):
		for file in csvfiles:
			currentDF = pd.read_csv(os.path.join(location, file))
			if substitute_dev == 'True':
				currentDF['Actual'] = np.where(pd.isnull(currentDF['Actual']),currentDF['Dev'],currentDF['Actual'])
				deviations[str(file[:-4])] = currentDF['Actual']
			else:
				deviations[str(file[:-4])] = currentDF['Actual']		
		deviations.to_csv(location + r'\ΩActuals.csv', index=False)

	return deviations

def update_master():
	#define color for rework rounds
	color_index = {0:colors.WHITE,1:colors.GREEN,2:colors.YELLOW,3:'FF6600'}

	#load the master and set it to the 'Data' tab
	wb = load_workbook(master,read_only=False, keep_vba=True)
	ws = wb['Data']

	#Element column in the deviations DF is only needed for the combine file
	del deviations['Element']


	if rw_round == 0:		
		#if its not rework then there is nothing to compare to
		#therefore the combined deviations (or actuals) is the masterDF
		masterDF = deviations

	else:

		#if were in rework then we need to create a dataframe from the current
		#master 'Data' sheet

		#load all the values from the "Data" sheet
		data = ws.values

		#get the columns (not including the Elements/SN column)
		cols = next(data)[1:]

		#convert data to a list
		data = list(data)

		#use the values in the first column (Elements/SN) to define the index 
		#(will use for length not values)
		idx = [r[0] for r in data if r[0] != None]

		#define the range containing the actual data by slicing each row
		#for the length of the index
		data = (islice(r,1,None) for r in data[:len(idx)])

		#construct the master dataframe using the gathered elements 
		#(data range,index column,column headers)
		masterDF = pd.DataFrame(data,index=idx,columns=cols)

		#in order to use the update function the indexes between the two 
		#dataframes must match. Reset index using defualt (0,1,2,...)
		#inplace changes the current dataframe rather than making a copy
		#drop removes the original index column rather than adding it to
		#the data range
		masterDF.reset_index(inplace=True,drop=True)

		#remove any empty columns
		del masterDF[None]
		
		#use the current round of deviations dataframe to
		#update the values in the masterDF where the 
		#columns and indexes match and values are different
		masterDF.update(deviations)

	#convert the dataframe to rows to update exisiting
	#excel file easier. index=False excludes the index column
	rows = dataframe_to_rows(masterDF,index=False)

	#iterate the columns and rows of the rows object
	#to update the excel master "Data" sheet
	for r_idx,row in enumerate(rows,1):
		for c_idx,value in enumerate(row,2):			
			cell = ws.cell(row=r_idx,column=c_idx)

			#try to convert to float for formatting
			try:
				value = float(value)
			except ValueError: #current row,column value is not a number
				pass

			cell.value = value

			#if in a rework round, change the header cell foreground to
			#a color based on the color_index dictionary and round number
			if rw_round > 0:
				if r_idx == 1 and value in deviations.columns:
					cell.fill = PatternFill(fgColor=color_index[rw_round],fill_type="solid")

	#save and close the master
	wb.save(master)
	wb.close()

#use registry to find install location of excel on current machine
def find_excel():
	handle = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
    r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\excel.exe")
	return winreg.EnumValue(handle, 0)[1]
	


if __name__ == "__main__":

	#make logger
	log = Logger().make_logger('Deviations')

	#get config info
	cfg = ConfigParser()
	configfile = r'config.ini'
	cfg.read(configfile)
	location = cfg.get('ScrapeData', 'location')
	rw_round = int(cfg.get('ScrapeData', 'rw_round'))
	substitute_dev = cfg.get('ScrapeData', 'substitute_dev')
	master = cfg.get('ScrapeData','master')
	base_location = get_location()
	csvfiles = get_csv_files()

	#create deviations dataframe and make combined file(optional)
	deviations = get_deviations()

	#update the master
	update_master()

	#find the excel.exe path
	excel = find_excel()

	#launch the master
	subprocess.Popen([excel,master])








